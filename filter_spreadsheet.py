#!/usr/bin/env python3
import argparse
import csv
import unicodedata
from pathlib import Path

def normalise_string(s):
    """Remove all whitespace and hidden characters, convert to lowercase"""
    if s is None:
        return ''
    s = str(s)
    # Remove all whitespace characters (spaces, tabs, newlines, etc.)
    s = ''.join(s.split())
    # Remove other control/hidden characters
    s = ''.join(char for char in s if unicodedata.category(char)[0] != 'C')
    # Convert to lowercase for case-insensitive comparison
    return s.lower()

def read_filter_values(txt_path, skip_header=False):
    """Read filter values from text file, returning both normalised set and original values mapping"""
    with open(txt_path, 'r', encoding='utf-8') as f:
        lines = [line.strip() for line in f if line.strip()]
        if skip_header and lines:
            lines = lines[1:]
        # Map normalised -> original for reporting unmatched values
        normalised_to_original = {normalise_string(line): line for line in lines}
        return set(normalised_to_original.keys()), normalised_to_original

def detect_delimiter(file_path):
    """Detect delimiter for CSV/TSV files"""
    suffix = Path(file_path).suffix.lower()
    if suffix == '.tsv':
        return '\t'
    elif suffix == '.csv':
        return ','
    else:
        # Try to detect from content
        with open(file_path, 'r', encoding='utf-8') as f:
            first_line = f.readline()
            if '\t' in first_line:
                return '\t'
            return ','

def filter_csv_tsv(input_path, output_path, filter_col, filter_values, exclude=False):
    """Filter CSV or TSV files"""
    delimiter = detect_delimiter(input_path)
    out_suffix = Path(output_path).suffix.lower()
    out_delimiter = '\t' if out_suffix == '.tsv' else ','
    
    matched_filter_values = set()
    
    with open(input_path, 'r', newline='', encoding='utf-8') as infile, \
         open(output_path, 'w', newline='', encoding='utf-8') as outfile:
        
        reader = csv.DictReader(infile, delimiter=delimiter)
        
        if filter_col not in reader.fieldnames:
            raise ValueError(f"Column '{filter_col}' not found. Available columns: {reader.fieldnames}")
        
        writer = csv.DictWriter(outfile, fieldnames=reader.fieldnames, delimiter=out_delimiter)
        writer.writeheader()
        
        matches = 0
        non_matches = 0
        for row in reader:
            normalised_val = normalise_string(row[filter_col])
            is_match = normalised_val in filter_values
            if is_match:
                matched_filter_values.add(normalised_val)
            
            # In exclude mode, keep non-matching rows; otherwise keep matching rows
            if is_match != exclude:
                writer.writerow(row)
                matches += 1
            else:
                non_matches += 1
    
    return matches, non_matches, matched_filter_values

def filter_excel(input_path, output_path, filter_col, filter_values, exclude=False):
    """Filter Excel files using openpyxl"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel files. Install with: pip install openpyxl")
    
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    
    # Get header row
    headers = [cell.value for cell in ws[1]]
    
    if filter_col not in headers:
        raise ValueError(f"Column '{filter_col}' not found. Available columns: {headers}")
    
    col_idx = headers.index(filter_col)
    
    # Create output workbook
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    
    # Write header
    out_ws.append(headers)
    
    # Filter rows
    matches = 0
    non_matches = 0
    matched_filter_values = set()
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        normalised_val = normalise_string(row[col_idx])
        is_match = normalised_val in filter_values
        if is_match:
            matched_filter_values.add(normalised_val)
        
        # In exclude mode, keep non-matching rows; otherwise keep matching rows
        if is_match != exclude:
            out_ws.append(row)
            matches += 1
        else:
            non_matches += 1
    
    # Determine output format
    out_suffix = Path(output_path).suffix.lower()
    if out_suffix in ['.csv', '.tsv']:
        # Save as CSV/TSV instead
        delimiter = '\t' if out_suffix == '.tsv' else ','
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f, delimiter=delimiter)
            for row in out_ws.iter_rows(values_only=True):
                writer.writerow(row)
    else:
        out_wb.save(output_path)
    
    return matches, non_matches, matched_filter_values

def main():
    parser = argparse.ArgumentParser(
        description='Filter CSV/TSV/XLSX rows based on values in a text file'
    )
    parser.add_argument('--input', required=True, help='Input file (CSV, TSV, or XLSX)')
    parser.add_argument('--output', required=True, help='Output file (CSV, TSV, or XLSX)')
    parser.add_argument('--txt', required=True, help='Text file with filter values (one per line)')
    parser.add_argument('--filter', required=True, help='Column name to filter on (e.g., "Process ID")')
    parser.add_argument('--skip-header', action='store_true', 
                        help='Skip first line of filter txt file (if it has a header)')
    parser.add_argument('--exclude', action='store_true',
                        help='Invert filtering: remove matching rows instead of keeping them')
    parser.add_argument('--unmatched-file', 
                        help='Optional: write unmatched filter values to this file')
    
    args = parser.parse_args()
    
    input_path = Path(args.input)
    output_path = Path(args.output)
    
    # Read filter values (get both normalised set and mapping to originals)
    filter_values, normalised_to_original = read_filter_values(args.txt, args.skip_header)
    print(f"Loaded {len(filter_values)} filter values from {args.txt}")
    if args.exclude:
        print("Mode: EXCLUDE (removing matching rows)")
    else:
        print("Mode: INCLUDE (keeping matching rows only)")
    
    # Determine input type and process
    suffix = input_path.suffix.lower()
    
    if suffix in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
        matches, non_matches, matched_filter_values = filter_excel(
            args.input, args.output, args.filter, filter_values, exclude=args.exclude
        )
    elif suffix in ['.csv', '.tsv', '.txt']:
        matches, non_matches, matched_filter_values = filter_csv_tsv(
            args.input, args.output, args.filter, filter_values, exclude=args.exclude
        )
    else:
        # Try CSV/TSV as fallback
        print(f"Unknown extension '{suffix}', attempting to parse as delimited text...")
        matches, non_matches, matched_filter_values = filter_csv_tsv(
            args.input, args.output, args.filter, filter_values, exclude=args.exclude
        )
    
    # Calculate unmatched filter values
    unmatched_normalised = filter_values - matched_filter_values
    unmatched_original = [normalised_to_original[n] for n in unmatched_normalised]
    
    # Print summary
    print(f"\nFiltering complete:")
    print(f"  Rows written to output:    {matches}")
    print(f"  Rows skipped:              {non_matches}")
    print(f"  Filter values matched:     {len(matched_filter_values)} / {len(filter_values)}")
    print(f"  Filter values not found:   {len(unmatched_original)}")
    
    # Print unmatched filter values
    if unmatched_original:
        print(f"\nFilter values not found in '{args.filter}' column:")
        for val in sorted(unmatched_original):
            print(f"  - {val}")
        
        # Optionally write to file
        if args.unmatched_file:
            with open(args.unmatched_file, 'w', encoding='utf-8') as f:
                for val in sorted(unmatched_original):
                    f.write(f"{val}\n")
            print(f"\nUnmatched values written to: {args.unmatched_file}")
    
    print(f"\nOutput written to {args.output}")

if __name__ == '__main__':
    main()
