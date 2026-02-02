#!/usr/bin/env python3
"""
Taxonomy File Splitter

Splits a CSV, TSV, or XLSX file into multiple files based on a specified taxonomic 
rank (phylum, class, order, family, genus, etc.).

Two modes of operation:
1. NCBI API mode (default): Queries NCBI Taxonomy database using taxid column
2. Taxonomy column mode (--use-taxonomy): Uses existing taxonomy columns in the file

Input:
    - CSV, TSV, or XLSX file (auto-detected)
    - NCBI mode: requires taxid in column 4
    - Taxonomy mode: requires column matching the specified --rank

Output:
    - Multiple files named: [input_basename]_[rank_value].[original_extension]
    - Output format matches input format (CSV -> CSV, TSV -> TSV, XLSX -> XLSX)

Usage:
    # NCBI API mode - queries NCBI for taxonomy info
    python taxonomy_splitter.py -i specimens.csv -e you@example.com
    python taxonomy_splitter.py -i specimens.tsv -e you@example.com -r order

    # Taxonomy column mode - uses existing columns in file
    python taxonomy_splitter.py -i specimens.tsv --use-taxonomy
    python taxonomy_splitter.py -i specimens.xlsx -t -r family

Arguments:
    -i, --input         Input file path (required)
    -e, --email         Email for NCBI API (required unless --use-taxonomy)
    -r, --rank          Taxonomic rank to split on (default: phylum)
    -t, --use-taxonomy  Use taxonomy columns from file instead of NCBI API

Requirements:
    - biopython (for NCBI API mode)
    - openpyxl (for XLSX support)
"""

import argparse
import csv
import os
import time
from Bio import Entrez

def parse_arguments():
    parser = argparse.ArgumentParser(
        description='Split CSV/TSV/XLSX file by taxonomic rank based on taxids.'
    )
    parser.add_argument('--input', '-i', required=True, help='Input file path (CSV, TSV, or XLSX)')
    parser.add_argument(
        '--email', '-e', 
        required=False,
        help='Email address for NCBI API access (required unless --use-taxonomy is set)'
    )
    parser.add_argument(
        '--rank', '-r', 
        default='phylum', 
        help='Taxonomic rank to split on (default: phylum). Case-insensitive.'
    )
    parser.add_argument(
        '--use-taxonomy', '-t',
        action='store_true',
        help='Use taxonomy columns from input file instead of querying NCBI. '
             'Requires a column matching the specified --rank.'
    )
    return parser.parse_args()


def detect_file_type(filepath):
    """Detect file type based on extension and content."""
    ext = os.path.splitext(filepath)[1].lower()
    
    if ext == '.xlsx':
        return 'xlsx', None
    
    # For text files, detect delimiter
    with open(filepath, 'r', newline='') as f:
        sample = f.read(8192)
        
    # Use csv.Sniffer to detect delimiter
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',\t')
        delimiter = dialect.delimiter
    except csv.Error:
        # Default to comma if detection fails
        delimiter = ','
    
    if delimiter == '\t' or ext == '.tsv':
        return 'tsv', '\t'
    else:
        return 'csv', ','


def read_input_file(filepath):
    """Read input file and return header and rows."""
    file_type, delimiter = detect_file_type(filepath)
    
    if file_type == 'xlsx':
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for XLSX support. Install with: pip install openpyxl")
        
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = list(rows[0])
        data_rows = [list(row) for row in rows[1:]]
        return header, data_rows, file_type, delimiter
    
    else:
        with open(filepath, 'r', newline='') as f:
            reader = csv.reader(f, delimiter=delimiter)
            header = next(reader)
            data_rows = [row for row in reader]
        return header, data_rows, file_type, delimiter


def write_output_file(filepath, header, rows, file_type, delimiter):
    """Write output file in the same format as input."""
    if file_type == 'xlsx':
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(header)
        for row in rows:
            ws.append(row)
        wb.save(filepath)
    else:
        with open(filepath, 'w', newline='') as f:
            writer = csv.writer(f, delimiter=delimiter)
            writer.writerow(header)
            writer.writerows(rows)


def get_extension(file_type):
    """Return appropriate file extension for file type."""
    if file_type == 'xlsx':
        return '.xlsx'
    elif file_type == 'tsv':
        return '.tsv'
    else:
        return '.csv'


def find_column_index(header, column_name):
    """Find column index by name (case-insensitive). Returns None if not found."""
    column_name_lower = column_name.lower()
    for i, col in enumerate(header):
        if str(col).lower() == column_name_lower:
            return i
    return None


def get_rank_for_taxid(taxid, email, target_rank):
    """Retrieve taxonomic rank value for a given taxid using NCBI's API."""
    Entrez.email = email
    target_rank_lower = target_rank.lower()
    
    try:
        handle = Entrez.efetch(db="taxonomy", id=str(taxid), retmode="xml")
        records = Entrez.read(handle)
        handle.close()
        
        if not records or len(records) == 0:
            print(f"Warning: No taxonomy record found for taxid {taxid}")
            return "Unknown"
        
        # Find the target rank in the lineage
        lineage = records[0].get('LineageEx', [])
        for entry in lineage:
            if entry.get('Rank', '').lower() == target_rank_lower:
                return entry.get('ScientificName', 'Unknown')
        
        # Check if the taxon itself is at the target rank
        if records[0].get('Rank', '').lower() == target_rank_lower:
            return records[0].get('ScientificName', 'Unknown')
        
        print(f"Warning: No {target_rank} found in lineage for taxid {taxid}")
        return "Unknown"
    
    except Exception as e:
        print(f"Error retrieving taxonomy for {taxid}: {str(e)}")
        return "Unknown"


def main():
    args = parse_arguments()
    
    input_file = args.input
    email = args.email
    target_rank = args.rank.lower()
    use_taxonomy = args.use_taxonomy
    
    # Validate email is provided when using NCBI API mode
    if not use_taxonomy and not email:
        print("Error: --email is required when not using --use-taxonomy mode.")
        return
    
    # Get base filename for output
    input_basename = os.path.splitext(os.path.basename(input_file))[0]
    output_dir = os.path.dirname(input_file) or '.'
    
    # Read the input file
    print(f"Reading input file: {input_file}")
    header, data_rows, file_type, delimiter = read_input_file(input_file)
    print(f"Detected file type: {file_type}" + (f" (delimiter: {'tab' if delimiter == chr(9) else repr(delimiter)})" if delimiter else ""))
    print(f"Splitting by rank: {target_rank}")
    
    # Determine mode and validate
    if use_taxonomy:
        print("Mode: Using taxonomy columns from file")
        rank_col_idx = find_column_index(header, target_rank)
        if rank_col_idx is None:
            print(f"Error: Column '{target_rank}' not found in header.")
            print(f"Available columns: {', '.join(str(h) for h in header)}")
            return
        print(f"Found '{target_rank}' column at index {rank_col_idx}")
    else:
        print("Mode: Querying NCBI Taxonomy API")
    
    # Group rows by taxonomic rank
    rank_data = {}
    rank_mapping = {}  # Cache taxid to rank value mapping (only used in API mode)
    
    for row in data_rows:
        if use_taxonomy:
            # Get rank value directly from column
            if len(row) <= rank_col_idx:
                print(f"Warning: Skipping row with insufficient columns: {row}")
                continue
            
            rank_value = str(row[rank_col_idx]).strip() if row[rank_col_idx] else "Unknown"
            if not rank_value:
                rank_value = "Unknown"
            
            row_id = row[0] if row else "unknown"
            
        else:
            # Original NCBI API mode
            if len(row) < 4:
                print(f"Warning: Skipping invalid row: {row}")
                continue
            
            taxid = row[3]
            row_id = row[0] if row else "unknown"
            
            # Skip empty taxids
            if not taxid or str(taxid).strip() == '':
                print(f"Warning: Skipping row with empty taxid: {row_id}")
                continue
            
            # Get rank value, using cache if available
            if taxid in rank_mapping:
                rank_value = rank_mapping[taxid]
            else:
                rank_value = get_rank_for_taxid(taxid, email, target_rank)
                rank_mapping[taxid] = rank_value
                # Be nice to NCBI's API with a short delay
                time.sleep(0.35)
        
        # Add row to the appropriate rank group
        if rank_value not in rank_data:
            rank_data[rank_value] = []
        rank_data[rank_value].append(row)
        
        print(f"Processed: {row_id} -> {target_rank}: {rank_value}")
    
    # Write output files for each rank value
    extension = get_extension(file_type)
    for rank_value, rows in rank_data.items():
        # Create safe filename
        safe_rank_value = "".join([c if c.isalnum() else "_" for c in str(rank_value)])
        output_file = os.path.join(output_dir, f"{input_basename}_{safe_rank_value}{extension}")
        
        write_output_file(output_file, header, rows, file_type, delimiter)
        
        print(f"Created file {output_file} with {len(rows)} entries")
    
    print(f"\nDone! Created {len(rank_data)} files.")


if __name__ == "__main__":
    main()
